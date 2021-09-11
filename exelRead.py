import openpyxl
import excelobject
import pdfsheetfunction
import cratepdf
from fpdf import FPDF
from openpyxl.descriptors.serialisable import Serialisable
createpdf=cratepdf
pattern= "{:.2f}"
# empd =excelobject.excelproperty
style=pdfsheetfunction

book = openpyxl.load_workbook('D:\sle.xlsx',data_only=True)
print(book.sheetnames)
sheet = book["PAYROLL"]
row_count = sheet.max_row
paysheetrows=[]
for x in range(4, row_count+1):
    empd =excelobject.excelproperty
    empd.CompanyName=sheet['A1'].value
    empd.payRollMonth=sheet['B2'].value
    empd.epfNo =sheet['A'+str(x)].value
    empd.name=sheet['B'+str(x)].value
    empd.rate=pattern.format(sheet['C'+str(x)].value)
    empd.days=sheet['D'+str(x)].value
    empd.amount=pattern.format(sheet['E'+str(x)].value)
    empd.ot1hrs=sheet['F'+str(x)].value
    empd.ot1rate=pattern.format(sheet['G'+str(x)].value)
    empd.ot1ttl=pattern.format(sheet['H'+str(x)].value)
    empd.ot2hrs=sheet['I'+str(x)].value
    empd.ot2rate=pattern.format(sheet['J'+str(x)].value)
    empd.ot2ttl=pattern.format(sheet['K'+str(x)].value)
    empd.total=pattern.format(sheet['L'+str(x)].value)
    empd.attendenceAllowence=pattern.format(sheet['M'+str(x)].value)
    empd.attendenceincentive=pattern.format(sheet['N'+str(x)].value)
    empd.epf8amount=pattern.format(sheet['O'+str(x)].value)
    empd.totalearning=pattern.format(sheet['P'+str(x)].value)
    empd.epf12amount=pattern.format(sheet['Q'+str(x)].value)
    empd.etf3=pattern.format(sheet['R'+str(x)].value)
    empd.grosssalary=pattern.format(sheet['S'+str(x)].value)
#     paysheetrows.append(empd)
    createpdf.createpdf(payroll=empd)

    


# a1 = sheet['A1']
# a2 = sheet['A2']
# a3 = sheet.cell(row=3, column=1)

# print(a1.value)
# print(a2.value) 
# print(a3.value)
# print(empd.epfNo)
# print(row_count)





